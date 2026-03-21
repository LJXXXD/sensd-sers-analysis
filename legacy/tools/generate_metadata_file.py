import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment


def generate_metadata_file(signal_folder, output_file):
    """
    Generates a metadata file from signal filenames in the specified folder.

    Parameters:
    - signal_folder (str): Path to the folder containing signal files.
    - output_file (str): Path to save the generated metadata file (Excel format).

    The function assumes filenames follow the format:
    [Sensor_Config]_[Sensor_ID]_[Test_ID]_[Serotype].xlsx
    """

    # Mapping for full serotype names
    serotype_mapping = {"ty": "Typhimurium", "en": "Enterica", "mix": "Ty+En"}

    metadata = []

    for file in os.listdir(signal_folder):
        if file.endswith(".xlsx"):
            # Extract components from the filename
            parts = file.split("_")
            if len(parts) != 4:
                print(f"Skipping file {file} due to incorrect format.")
                continue

            # Remove the first letter from IDs and convert Sensor ID and Test ID to integers
            sensor_config = parts[0][1:]  # Remove 'C'
            sensor_id = int(parts[1][1:])  # Remove 'S' and convert to integer
            test_id = int(parts[2][1:])  # Remove 'T' and convert to integer
            serotype = (
                parts[3].replace(".xlsx", "").lower()
            )  # Convert to lowercase for case-insensitive matching

            # Map serotype abbreviation to full name
            serotype_full = serotype_mapping.get(serotype, serotype)

            metadata.append(
                {
                    "Filename": file,
                    "Config ID": sensor_config,
                    "Sensor ID": sensor_id,  # Keep as integer for numeric sorting
                    "Test ID": test_id,  # Keep as integer for numeric sorting
                    "Serotype": serotype_full,
                    "Sensing Date": None,  # Placeholder for sensing date
                    "Sample Preparation Date": None,  # Placeholder for preparation date
                    "Sample Preserving Temperature": None,  # Placeholder for temperature
                    "Notes": None,  # Placeholder for notes
                    "Tester": None,  # Placeholder for tester name
                }
            )

    # Convert to DataFrame
    metadata_df = pd.DataFrame(metadata)

    # Sort by Config ID, Sensor ID (numeric), and Test ID (numeric)
    metadata_df = metadata_df.sort_values(by=["Config ID", "Sensor ID", "Test ID"], ascending=True)

    # Convert Sensor ID and Test ID back to strings for Excel formatting
    metadata_df["Sensor ID"] = metadata_df["Sensor ID"].astype(str)
    metadata_df["Test ID"] = metadata_df["Test ID"].astype(str)

    # Replace None values with empty strings
    metadata_df = metadata_df.where(pd.notnull(metadata_df), "")

    # Save to Excel with text formatting for all cells
    workbook = Workbook()
    sheet = workbook.active

    # Define text style with center alignment
    text_style = NamedStyle(name="text_style", number_format="@")
    text_style.font = Font(size=16)  # Set font size to 16
    text_style.alignment = Alignment(horizontal="center", vertical="center")  # Center-align text

    if "text_style" not in workbook.named_styles:
        workbook.add_named_style(text_style)

    # Write header row
    sheet.append(metadata_df.columns.tolist())

    # Write data rows
    for row in metadata_df.itertuples(index=False):
        sheet.append(
            [cell if cell is not None else "" for cell in row]
        )  # Leave empty cells instead of "None"

    # Apply text style and calculate column widths
    column_widths = [0] * len(metadata_df.columns)  # Initialize column widths
    for row_idx, row in enumerate(sheet.iter_rows()):
        for col_idx, cell in enumerate(row):
            cell.style = text_style  # Apply text style
            # Update column width to fit the longest text
            column_widths[col_idx] = max(
                column_widths[col_idx], len(str(cell.value) if cell.value else "")
            )

    # Set column widths
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = (
            width + 4
        )  # Add padding for readability

    # Save the workbook
    workbook.save(output_file)
    print(f"Metadata file saved to {output_file} with centered text and adjusted column widths.")


# Example usage
signal_folder = "./data/SERS Data 6 (Dec 2024)/December Signals"  # Replace with your folder path
output_file = "./data/SERS Data 6 (Dec 2024)/December Signals/_metadata.xlsx"
generate_metadata_file(signal_folder, output_file)
