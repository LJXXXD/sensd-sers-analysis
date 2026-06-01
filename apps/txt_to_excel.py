"""
Instrument TXT to Excel merger — Streamlit prep utility.

Converts raw Raman instrument ``.txt`` exports into embedded-metadata Excel
workbooks for the SERS analysis tool. Lives entirely under ``apps/``.
"""

from __future__ import annotations

import json
import logging
import re
import uuid
from datetime import date, datetime, time
from io import BytesIO, StringIO
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd
import streamlit as st

from components.shared_ui import render_figure_stretch
from sensd_sers_analysis.visualization import plot_spectra

if TYPE_CHECKING:
    from streamlit.delta_generator import DeltaGenerator

logger = logging.getLogger(__name__)

APP_MODE_KEY = "_app_mode"
APP_MODE_ANALYSIS = "analysis"
APP_MODE_PREP = "prep"

MERGED_PREVIEW_KEY = "_txt2excel_merged_preview"
EXPORT_BYTES_KEY = "_txt2excel_export_bytes"
EXPORT_FILENAME_KEY = "_txt2excel_export_filename"
PREP_UPLOADER_RESET_KEY = "_txt2excel_uploader_reset"
TEMPLATE_IMPORT_UPLOADER_RESET_KEY = "_txt2excel_template_import_reset"
TEMPLATE_IMPORT_FEEDBACK_KEY = "_txt2excel_template_import_feedback"
TEMPLATE_IMPORT_PENDING_VALUES_KEY = "_txt2excel_template_import_pending_values"
TEMPLATE_EXPORT_SELECT_ALL_KEY = "txt2excel_template_export_select_all"
TEMPLATE_EXPORT_FILENAME = "SERS_metadata_preset.json"
TEMPLATE_VERSION = 1
DEFAULT_MIN_SHIFT = 560.9
MERGED_PREVIEW_FIGSIZE = (10.0, 6.0)
MERGED_PREVIEW_HUE_COL = "concentration_cfu_ml"
MERGED_PREVIEW_LEGEND_TITLE = "Concentrations (CFU/mL)"

TARGET_CONCENTRATION_LABEL = "Target Concentration (CFU/mL)"
ACTUAL_CONCENTRATION_LABEL = "Actual Concentration (CFU/mL)"
FILE_NAME_LABEL = "File Name"
SPECIAL_TREATMENT_LABEL = "Special Treatment"
PREP_TARGET_CONCENTRATION_HEADER = "Target Concentration"
PREP_ACTUAL_CONCENTRATION_HEADER = "Actual Concentration"
PREP_SPECIAL_TREATMENT_HEADER = "Special Treatment"
REQUIRED_FIELD_LABEL_SUFFIX = " *"
OPTIONAL_FIELD_LABEL_SUFFIX = " (optional)"
INTENSITY_HEADER = "Relative Light intensity (a.u)"
RAMAN_SHIFT_HEADER = "Raman Shift"

# Metadata field numbers whose column-B values must be numeric in the workbook.
NUMERIC_METADATA_FIELD_NUMBERS = frozenset({1, 2, 3, 4, 6, 7})
DATE_METADATA_FIELD_NUMBERS = frozenset({13})
TIME_METADATA_FIELD_NUMBERS = frozenset({14})
OPTIONAL_METADATA_FIELD_NUMBERS = frozenset({16})
METADATA_FIELD_SPECS: tuple[tuple[int, str, str], ...] = (
    (1, "Disk Diameter (nm)", "txt2excel_meta_disk_diameter_nm"),
    (2, "Periodicity (µm)", "txt2excel_meta_periodicity_um"),
    (3, "Thickness (nm)", "txt2excel_meta_thickness_nm"),
    (4, "Core Diameter (µm)", "txt2excel_meta_core_diameter_um"),
    (5, "Sensor Model", "txt2excel_meta_sensor_model"),
    (6, "Integration Time (ms)", "txt2excel_meta_integration_time_ms"),
    (7, "Scan Average", "txt2excel_meta_scan_average"),
    (8, "Sensor ID", "txt2excel_meta_sensor_id"),
    (9, "Test ID", "txt2excel_meta_test_id"),
    (10, "Connection ID", "txt2excel_meta_connection_id"),
    (11, "Serotype", "txt2excel_meta_serotype"),
    (12, "Rinsate Type", "txt2excel_meta_rinsate_type"),
    (13, "Date", "txt2excel_meta_date"),
    (14, "Testing Time", "txt2excel_meta_testing_time"),
    (15, "Operator", "txt2excel_meta_operator"),
    (16, "Notes", "txt2excel_meta_notes"),
)
# Logical groups (5 + 2 + 3 + 2 + 3 + notes); Excel dividers follow the last field in each.
METADATA_LOGICAL_GROUPS: tuple[tuple[int, ...], ...] = (
    (1, 2, 3, 4, 5),
    (6, 7),
    (8, 9, 10),
    (11, 12),
    (13, 14, 15),
    (16,),
)
METADATA_UI_COLUMN_COUNT = 5
# Three five-column UI rows (15 fields) before full-width notes.
METADATA_UI_ROWS: tuple[tuple[int, ...], ...] = (
    (1, 2, 3, 4, 5),
    (6, 7, 8, 9, 10),
    (11, 12, 13, 14, 15),
)
_PREP_LAYOUT_STABILITY_CSS_KEY = "_txt2excel_prep_layout_css_injected"
METADATA_WIDGET_KEYS = frozenset(widget_key for _, _, widget_key in METADATA_FIELD_SPECS)
RELOAD_CLEAR_WIDGET_KEYS = frozenset(
    {
        "txt2excel_meta_sensor_id",
        "txt2excel_meta_test_id",
        "txt2excel_meta_connection_id",
        "txt2excel_meta_serotype",
        "txt2excel_meta_testing_time",
        "txt2excel_meta_notes",
    }
)
RELOAD_PERSIST_WIDGET_KEYS = frozenset(METADATA_WIDGET_KEYS - RELOAD_CLEAR_WIDGET_KEYS)
PERSISTENT_METADATA_SNAPSHOT_KEY = "_txt2excel_persistent_metadata"
RESTORE_METADATA_AFTER_RELOAD_KEY = "_txt2excel_restore_metadata_after_reload"


def enter_prep_mode() -> None:
    """Switch the app shell to the TXT-to-Excel prep utility."""

    st.session_state[APP_MODE_KEY] = APP_MODE_PREP


def enter_analysis_mode() -> None:
    """Return to the main SERS analysis explorer."""

    st.session_state[APP_MODE_KEY] = APP_MODE_ANALYSIS
    st.session_state.pop(MERGED_PREVIEW_KEY, None)
    st.session_state.pop(EXPORT_BYTES_KEY, None)
    st.session_state.pop(EXPORT_FILENAME_KEY, None)


def _clear_session_keys_by_prefix(prefix: str) -> None:
    """Remove all session-state keys that start with ``prefix``."""

    for key in list(st.session_state.keys()):
        if key.startswith(prefix):
            del st.session_state[key]


def _merge_widget_values_into_snapshot(
    snapshot: dict[str, str],
    widget_values: dict[str, str],
    *,
    persist_keys: frozenset[str] = RELOAD_PERSIST_WIDGET_KEYS,
) -> dict[str, str]:
    """
    Merge widget values into a persistent metadata snapshot.

    Streamlit drops widget session keys when inputs are not rendered (e.g. after
    Reload Data clears uploads). The snapshot survives across those runs.
    """

    updated = dict(snapshot)
    for key in persist_keys:
        if key in widget_values:
            updated[key] = widget_values[key]
    return updated


def _clear_reload_fields_in_snapshot(
    snapshot: dict[str, str],
    *,
    clear_keys: frozenset[str] = RELOAD_CLEAR_WIDGET_KEYS,
) -> dict[str, str]:
    """Blank run-specific metadata keys in the persistent snapshot."""

    updated = dict(snapshot)
    for key in clear_keys:
        updated[key] = ""
    return updated


def _sync_persistent_metadata_snapshot(widget_values: dict[str, str] | None = None) -> None:
    """Save persistent metadata widget values into the cross-run snapshot."""

    values = (
        widget_values if widget_values is not None else _collect_metadata_values_by_widget_key()
    )
    snapshot = st.session_state.get(PERSISTENT_METADATA_SNAPSHOT_KEY, {})
    st.session_state[PERSISTENT_METADATA_SNAPSHOT_KEY] = _merge_widget_values_into_snapshot(
        snapshot,
        values,
    )


def _restore_persistent_metadata_widgets() -> None:
    """Restore persistent metadata widgets from the snapshot before rendering inputs."""

    snapshot = st.session_state.get(PERSISTENT_METADATA_SNAPSHOT_KEY, {})
    for field_number, _, key in METADATA_FIELD_SPECS:
        if key in RELOAD_PERSIST_WIDGET_KEYS and key in snapshot:
            st.session_state[key] = _coerce_metadata_widget_state_value(field_number, snapshot[key])


def _apply_pending_template_import_values() -> None:
    """Apply deferred template-import values before metadata widgets are rendered."""

    pending_values = st.session_state.pop(TEMPLATE_IMPORT_PENDING_VALUES_KEY, None)
    if not isinstance(pending_values, dict):
        return
    for widget_key, value in pending_values.items():
        if widget_key not in METADATA_WIDGET_KEYS:
            continue
        field_number = _field_number_for_widget_key(widget_key)
        if field_number is None:
            continue
        st.session_state[widget_key] = _coerce_metadata_widget_state_value(field_number, value)
    _sync_persistent_metadata_snapshot(
        {
            widget_key: _metadata_widget_state_to_string(
                field_number, st.session_state.get(widget_key)
            )
            for widget_key in pending_values
            if widget_key in METADATA_WIDGET_KEYS
            for field_number in [_field_number_for_widget_key(widget_key)]
            if field_number is not None
        }
    )


def clear_prep_uploads() -> None:
    """
    Reset uploaded TXT files and derived export artifacts.

    Instrument metadata (fields 1–7), date, operator, and rinsate type persist;
    run-specific fields, concentrations, and Raman bounds reset for the next batch.
    """

    logger.info("Clearing prep upload state (Reload Data clicked)")
    widget_values = {
        widget_key: _metadata_widget_state_to_string(
            field_number,
            st.session_state.get(widget_key),
        )
        for field_number, _, widget_key in METADATA_FIELD_SPECS
        if widget_key in st.session_state
    }
    snapshot = st.session_state.get(PERSISTENT_METADATA_SNAPSHOT_KEY, {})
    snapshot = _merge_widget_values_into_snapshot(snapshot, widget_values)
    snapshot = _clear_reload_fields_in_snapshot(snapshot)
    st.session_state[PERSISTENT_METADATA_SNAPSHOT_KEY] = snapshot

    st.session_state.pop(MERGED_PREVIEW_KEY, None)
    st.session_state.pop(EXPORT_BYTES_KEY, None)
    st.session_state.pop(EXPORT_FILENAME_KEY, None)
    st.session_state.pop("txt2excel_file_signature", None)
    st.session_state.pop("txt2excel_min_shift", None)
    st.session_state.pop("txt2excel_max_shift", None)
    for widget_key in RELOAD_CLEAR_WIDGET_KEYS:
        field_number = _field_number_for_widget_key(widget_key)
        if field_number in DATE_METADATA_FIELD_NUMBERS | TIME_METADATA_FIELD_NUMBERS:
            st.session_state[widget_key] = None
        else:
            st.session_state[widget_key] = ""
    _clear_session_keys_by_prefix("txt2excel_target_")
    _clear_session_keys_by_prefix("txt2excel_actual_")
    _clear_session_keys_by_prefix("txt2excel_treatment_")
    st.session_state[RESTORE_METADATA_AFTER_RELOAD_KEY] = True
    st.session_state[PREP_UPLOADER_RESET_KEY] = str(uuid.uuid4())


def render_prep_entry_in_sidebar(sidebar: DeltaGenerator) -> None:
    """
    Render a compact link-style control above the analysis data-loading section.

    Parameters
    ----------
    sidebar:
        Streamlit sidebar container.
    """

    sidebar.button(
        "Convert TXT → Excel",
        key="enter_txt_to_excel_mode",
        on_click=enter_prep_mode,
        use_container_width=True,
    )
    sidebar.markdown("---")


def _extract_cfu_sort_key(filename: str) -> int:
    """Extract the first integer in a filename for natural dilution ordering."""

    match = re.search(r"(\d+)", filename)
    return int(match.group(1)) if match else int(1e9)


def _parse_txt_content(content: str) -> pd.DataFrame:
    """
    Parse one instrument TXT export into Raman shift and intensity columns.

    Parameters
    ----------
    content:
        UTF-8 text of the instrument export.

    Returns
    -------
    pd.DataFrame
        Columns ``RamanShift`` and ``Value``.
    """

    df = pd.read_csv(
        StringIO(content),
        sep="\t",
        comment=">",
        header=None,
        names=["RamanShift", "Value"],
    )
    df["RamanShift"] = pd.to_numeric(df["RamanShift"], errors="coerce")
    return df.dropna(subset=["RamanShift", "Value"]).reset_index(drop=True)


def _validate_common_shift(
    file_contents: dict[str, str],
    *,
    atol: float = 1e-4,
) -> tuple[np.ndarray | None, str | None]:
    """
    Verify all uploaded TXT files share the same Raman shift grid.

    Returns
    -------
    tuple
        ``(common_shift, error_message)`` — shift array when valid, else ``None`` and message.
    """

    common_shift: np.ndarray | None = None
    for name, content in file_contents.items():
        shifts = _parse_txt_content(content)["RamanShift"].to_numpy()
        if common_shift is None:
            common_shift = shifts
            continue
        if not np.allclose(common_shift, shifts, atol=atol):
            return None, f"Raman shift mismatch in file: {name}"
    return common_shift, None


def _merge_txt_spectra(
    file_contents: dict[str, str],
    *,
    min_shift: float | None,
    max_shift: float | None,
) -> tuple[np.ndarray | None, list[np.ndarray] | None, str | None]:
    """
    Merge parsed TXT spectra into aligned Raman shift and intensity columns.

    Parameters
    ----------
    file_contents:
        Mapping of filename to UTF-8 file text, in column order.
    min_shift:
        Optional lower Raman shift bound (cm⁻¹).
    max_shift:
        Optional upper Raman shift bound (cm⁻¹).

    Returns
    -------
    tuple
        ``(raman_shift, intensity_columns, error_message)``.
    """

    names = list(file_contents.keys())
    if not names:
        return None, None, "No TXT files provided."

    common_shift, shift_error = _validate_common_shift(file_contents)
    if shift_error:
        return None, None, shift_error
    if common_shift is None:
        return None, None, "No Raman shift data found."

    min_final = float(min_shift) if min_shift is not None else float(common_shift.min())
    max_final = float(max_shift) if max_shift is not None else float(common_shift.max())
    if min_final >= max_final:
        return None, None, "Min Raman Shift must be less than Max."

    mask = (common_shift >= min_final) & (common_shift <= max_final)
    raman_shift = common_shift[mask]
    intensity_columns: list[np.ndarray] = []

    for name in names:
        df = _parse_txt_content(file_contents[name])
        trimmed = df.loc[mask, "Value"].to_numpy(dtype=float)
        if len(trimmed) != len(raman_shift):
            return None, None, f"Truncated row count mismatch for file: {name}"
        intensity_columns.append(trimmed)

    return raman_shift, intensity_columns, None


def _parse_required_number(raw: str) -> float | None:
    """Parse a required numeric field; returns ``None`` when empty or non-numeric."""

    stripped = raw.strip()
    if not stripped:
        return None
    try:
        return float(stripped)
    except ValueError:
        return None


def _parse_optional_number(raw: str) -> float | None:
    """
    Parse an optional numeric field.

    Returns ``None`` when empty (allowed). Returns ``None`` when non-numeric
    (caller should treat as validation failure).
    """

    return _parse_required_number(raw)


def _to_excel_number(value: float) -> int | float:
    """Coerce a numeric value for Excel cells (integers stored without decimals)."""

    if value.is_integer():
        return int(value)
    return value


def _parse_metadata_date(raw: str) -> date | None:
    """Parse a metadata date string in ISO or common US format."""

    stripped = raw.strip()
    if not stripped:
        return None
    try:
        return date.fromisoformat(stripped)
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(stripped, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_ampm_period(raw_period: str) -> str:
    """Return ``AM`` or ``PM`` from a case-insensitive period token."""

    token = raw_period.upper().replace(".", "")
    return "AM" if token.startswith("A") else "PM"


def _ampm_parse_candidates(raw: str) -> list[str]:
    """
    Build normalized 12-hour time strings with an explicit AM/PM suffix.

    Accepts compact input such as ``230pm`` or ``2:30 PM`` and expands it into
    ``strptime``-compatible candidates.
    """

    text = re.sub(r"\s+", " ", raw.strip())
    if not text:
        return []

    ampm_match = re.search(r"(a\.?m\.?|p\.?m\.?)\.?\s*$", text, re.IGNORECASE)
    if not ampm_match:
        return [text]

    period = _normalize_ampm_period(ampm_match.group(1))
    time_part = text[: ampm_match.start()].strip()
    if not time_part:
        return [f"12 {period}"]

    candidates: list[str] = []
    if ":" in time_part:
        candidates.append(f"{time_part} {period}")
    else:
        digits = re.sub(r"\D", "", time_part)
        if not digits:
            return []
        if len(digits) <= 2:
            candidates.append(f"{int(digits)} {period}")
        elif len(digits) == 3:
            candidates.append(f"{digits[0]}:{digits[1:]} {period}")
        else:
            hour_digits = digits[:2]
            minute_digits = digits[2:4]
            if int(hour_digits) > 12:
                candidates.append(f"{digits[0]}:{digits[1:3]} {period}")
            else:
                candidates.append(f"{int(hour_digits)}:{minute_digits} {period}")
                if len(digits) >= 4:
                    candidates.append(f"{int(digits[0])}:{digits[1:3]} {period}")

    deduped: list[str] = []
    for candidate in candidates:
        if candidate not in deduped:
            deduped.append(candidate)
    return deduped


def _parse_metadata_time(raw: str) -> time | None:
    """Parse a metadata time string in 12-hour AM/PM or legacy 24-hour format."""

    stripped = raw.strip()
    if not stripped:
        return None

    ampm_formats = ("%I:%M:%S %p", "%I:%M %p", "%I %p")
    for candidate in _ampm_parse_candidates(stripped):
        for fmt in ampm_formats:
            if fmt == "%I %p" and ":" in candidate:
                continue
            if fmt.startswith("%I:%M") and candidate.count(":") != fmt.count(":"):
                continue
            try:
                return datetime.strptime(candidate, fmt).time()
            except ValueError:
                continue

    legacy_formats = ("%H:%M:%S", "%H:%M")
    for candidate in (stripped,):
        for fmt in legacy_formats:
            if fmt == "%H:%M" and len(candidate) != 5:
                continue
            try:
                return datetime.strptime(candidate, fmt).time()
            except ValueError:
                continue
    return None


def _format_metadata_time_display(parsed: time) -> str:
    """Format a ``time`` value as ``HH:MM AM/PM`` with zero-padded hour and minute."""

    hour = parsed.hour % 12 or 12
    ampm = "AM" if parsed.hour < 12 else "PM"
    if parsed.second or parsed.microsecond:
        return f"{hour:02d}:{parsed.minute:02d}:{parsed.second:02d} {ampm}"
    return f"{hour:02d}:{parsed.minute:02d} {ampm}"


def _format_metadata_date(value: date | str | None) -> str:
    """Return a normalized ``YYYY-MM-DD`` string, or empty when invalid."""

    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    parsed = _parse_metadata_date(str(value))
    return parsed.isoformat() if parsed else ""


def _format_metadata_time(value: time | str | None) -> str:
    """Return a normalized 12-hour AM/PM time string, or empty when invalid."""

    if value is None:
        return ""
    if isinstance(value, time):
        parsed = value
    else:
        parsed = _parse_metadata_time(str(value))
    if parsed is None:
        return ""
    return _format_metadata_time_display(parsed)


def _coerce_metadata_widget_state_value(field_number: int, raw_value: Any) -> Any:
    """Convert stored template or snapshot text to a widget-ready session value."""

    if field_number in DATE_METADATA_FIELD_NUMBERS:
        if isinstance(raw_value, date):
            return raw_value
        if raw_value is None:
            return None
        stripped = str(raw_value).strip()
        return _parse_metadata_date(stripped) if stripped else None
    if field_number in TIME_METADATA_FIELD_NUMBERS:
        if isinstance(raw_value, time):
            return raw_value
        if raw_value is None:
            return None
        stripped = str(raw_value).strip()
        return _parse_metadata_time(stripped) if stripped else None
    if raw_value is None:
        return ""
    return str(raw_value).strip()


def _metadata_widget_state_to_string(field_number: int, raw_value: Any) -> str:
    """Serialize a metadata widget session value to workbook/template text."""

    if field_number in DATE_METADATA_FIELD_NUMBERS:
        return _format_metadata_date(raw_value)
    if field_number in TIME_METADATA_FIELD_NUMBERS:
        return _format_metadata_time(raw_value)
    if raw_value is None:
        return ""
    return str(raw_value).strip()


def _is_metadata_widget_value_empty(field_number: int, raw_value: Any) -> bool:
    """Return whether a metadata widget has no usable value."""

    return not _metadata_widget_state_to_string(field_number, raw_value)


def _collect_metadata_values() -> dict[str, str]:
    """Read metadata field values from Streamlit session state."""

    return {
        excel_label: _metadata_widget_state_to_string(
            field_number,
            st.session_state.get(widget_key),
        )
        for field_number, excel_label, widget_key in METADATA_FIELD_SPECS
    }


def _collect_metadata_values_by_widget_key(
    field_values: dict[str, str] | None = None,
) -> dict[str, str]:
    """
    Read metadata widget values from an explicit mapping or Streamlit session state.

    Parameters
    ----------
    field_values:
        Optional mapping of widget key to raw string. When omitted, session state is used.
    """

    if field_values is None:
        return {
            widget_key: _metadata_widget_state_to_string(
                field_number,
                st.session_state.get(widget_key),
            )
            for field_number, _, widget_key in METADATA_FIELD_SPECS
        }
    return {
        widget_key: _metadata_widget_state_to_string(field_number, field_values.get(widget_key))
        for field_number, _, widget_key in METADATA_FIELD_SPECS
    }


def _field_number_for_widget_key(widget_key: str) -> int | None:
    """Return the metadata field number for a widget key, if known."""

    return next(
        (field_number for field_number, _, key in METADATA_FIELD_SPECS if key == widget_key),
        None,
    )


def _is_template_field_exportable(field_number: int, raw_value: str) -> bool:
    """
    Return whether a metadata field has a valid, exportable value.

    Numeric fields (1–6) must parse as numbers. Date (12) and Testing Time (13) must use
    valid ``YYYY-MM-DD`` and 12-hour AM/PM formats. Optional Notes (16) may be omitted
    when empty. All other fields must be non-empty text.
    """

    stripped = raw_value.strip() if isinstance(raw_value, str) else raw_value
    if field_number in OPTIONAL_METADATA_FIELD_NUMBERS:
        if _is_metadata_widget_value_empty(field_number, stripped):
            return False
        return True
    if field_number in NUMERIC_METADATA_FIELD_NUMBERS:
        return _parse_required_number(str(stripped).strip()) is not None
    if field_number in DATE_METADATA_FIELD_NUMBERS:
        return _format_metadata_date(stripped) != ""
    if field_number in TIME_METADATA_FIELD_NUMBERS:
        return _format_metadata_time(stripped) != ""
    return bool(str(stripped).strip())


def _serialize_template_field_value(
    field_number: int, raw_value: str | Any
) -> str | int | float | None:
    """Convert a raw widget value to a JSON-safe template value, or ``None`` when invalid."""

    if field_number in NUMERIC_METADATA_FIELD_NUMBERS:
        parsed = _parse_required_number(str(raw_value).strip())
        if parsed is None:
            return None
        return _to_excel_number(parsed)
    if field_number in DATE_METADATA_FIELD_NUMBERS:
        formatted = _format_metadata_date(raw_value)
        return formatted if formatted else None
    if field_number in TIME_METADATA_FIELD_NUMBERS:
        formatted = _format_metadata_time(raw_value)
        return formatted if formatted else None
    stripped = str(raw_value).strip()
    if not stripped:
        return None
    return stripped


def _build_template_export_payload(
    field_values: dict[str, str],
    selected_keys: frozenset[str],
) -> dict[str, Any] | None:
    """
    Build a setup-template JSON payload from selected metadata widget keys.

    Only checked fields with valid values are included.
    """

    export_fields: dict[str, str | int | float] = {}
    for field_number, _, widget_key in METADATA_FIELD_SPECS:
        if widget_key not in selected_keys:
            continue
        serialized = _serialize_template_field_value(field_number, field_values.get(widget_key, ""))
        if serialized is None:
            continue
        export_fields[widget_key] = serialized
    if not export_fields:
        return None
    return {"version": TEMPLATE_VERSION, "fields": export_fields}


def _apply_template_import_to_values(
    payload: dict[str, Any],
    field_values: dict[str, str],
) -> tuple[dict[str, str], list[str]]:
    """
    Merge template payload values into a widget-key mapping.

    Returns
    -------
    tuple
        Updated values and human-readable warning messages.
    """

    warnings: list[str] = []
    if payload.get("version") != TEMPLATE_VERSION:
        warnings.append(f"Unsupported template version: {payload.get('version')!r}")
        return field_values, warnings

    raw_fields = payload.get("fields")
    if not isinstance(raw_fields, dict):
        warnings.append("Template is missing a valid 'fields' object.")
        return field_values, warnings

    updated = dict(field_values)
    for widget_key, raw_value in raw_fields.items():
        if widget_key not in METADATA_WIDGET_KEYS:
            warnings.append(f"Unknown template field ignored: {widget_key}")
            continue
        field_number = _field_number_for_widget_key(widget_key)
        if field_number is None:
            continue
        if isinstance(raw_value, bool):
            warnings.append(f"Invalid value for {widget_key}; skipped.")
            continue
        if isinstance(raw_value, (int, float)):
            if field_number in NUMERIC_METADATA_FIELD_NUMBERS:
                updated[widget_key] = str(_to_excel_number(float(raw_value)))
            else:
                updated[widget_key] = str(raw_value)
            continue
        if isinstance(raw_value, str):
            if field_number in NUMERIC_METADATA_FIELD_NUMBERS:
                parsed = _parse_required_number(raw_value)
                if parsed is None:
                    warnings.append(f"Invalid numeric value for {widget_key}; skipped.")
                    continue
                updated[widget_key] = str(_to_excel_number(parsed))
            elif field_number in DATE_METADATA_FIELD_NUMBERS:
                formatted = _format_metadata_date(raw_value)
                if not formatted:
                    warnings.append(f"Invalid date value for {widget_key}; skipped.")
                    continue
                updated[widget_key] = formatted
            elif field_number in TIME_METADATA_FIELD_NUMBERS:
                formatted = _format_metadata_time(raw_value)
                if not formatted:
                    warnings.append(f"Invalid time value for {widget_key}; skipped.")
                    continue
                updated[widget_key] = formatted
            else:
                updated[widget_key] = raw_value.strip()
            continue
        warnings.append(f"Unsupported value type for {widget_key}; skipped.")

    return updated, warnings


def _template_selection_key(widget_key: str) -> str:
    """Session-state key for a setup-template export checkbox."""

    return f"txt2excel_template_sel_{widget_key}"


def _all_template_export_fields_selected() -> bool:
    """Return whether every metadata export checkbox is selected."""

    return all(
        st.session_state.get(_template_selection_key(widget_key), False)
        for _, _, widget_key in METADATA_FIELD_SPECS
    )


def _apply_template_export_select_all() -> None:
    """Mirror the master select-all checkbox to every export field."""

    select_all = st.session_state.get(TEMPLATE_EXPORT_SELECT_ALL_KEY, False)
    for _, _, widget_key in METADATA_FIELD_SPECS:
        st.session_state[_template_selection_key(widget_key)] = select_all


def _validate_metadata() -> str | None:
    """Return an error message when any required metadata field is empty or invalid."""

    missing: list[str] = []
    invalid: list[str] = []
    for field_number, label, widget_key in METADATA_FIELD_SPECS:
        if field_number in OPTIONAL_METADATA_FIELD_NUMBERS:
            continue
        raw_value = st.session_state.get(widget_key)
        if _is_metadata_widget_value_empty(field_number, raw_value):
            missing.append(label)
            continue
        if field_number in DATE_METADATA_FIELD_NUMBERS and not _format_metadata_date(raw_value):
            invalid.append(f"{field_number}. {label} must be a valid date (YYYY-MM-DD).")
        elif field_number in TIME_METADATA_FIELD_NUMBERS and not _format_metadata_time(raw_value):
            invalid.append(f"{field_number}. {label} must be a valid time (e.g. 01:30 PM).")
    if missing:
        return f"Required metadata fields are missing: {', '.join(missing)}"
    if invalid:
        return invalid[0]
    return None


def _prep_field_label(text: str, *, optional: bool = False) -> str:
    """Format a prep UI label with required (``*``) or optional suffix."""

    suffix = OPTIONAL_FIELD_LABEL_SUFFIX if optional else REQUIRED_FIELD_LABEL_SUFFIX
    return f"{text}{suffix}"


def _prep_column_header_markup(text: str, *, optional: bool = False) -> str:
    """
    Format a column header for ``st.markdown`` with bold text.

    The required asterisk or optional suffix is placed outside the bold span so
    markdown parsing does not treat ``*`` as emphasis delimiters.
    """

    suffix = OPTIONAL_FIELD_LABEL_SUFFIX if optional else REQUIRED_FIELD_LABEL_SUFFIX
    return f"**{text}**{suffix}"


def _metadata_field_label(field_number: int, excel_label: str) -> str:
    """Format a numbered metadata widget label with required/optional suffix."""

    base = f"{field_number}. {excel_label}"
    if field_number in OPTIONAL_METADATA_FIELD_NUMBERS:
        return _prep_field_label(base, optional=True)
    return _prep_field_label(base, optional=False)


def _validate_prep_inputs(
    sorted_files: list[Any],
    target_inputs: list[str],
    actual_inputs: list[str],
    *,
    special_treatment_inputs: list[str] | None = None,
) -> tuple[
    dict[str, Any],
    list[int | float | str],
    list[int | float],
    list[str],
    list[str],
    str | None,
]:
    """
    Validate all user inputs before workbook export.

    Returns
    -------
    tuple
        ``(metadata, target_concentrations, actual_concentrations, source_txt_filenames,
        special_treatments, error_message)``.
    """

    metadata_error = _validate_metadata()
    if metadata_error:
        return {}, [], [], [], [], metadata_error

    metadata: dict[str, Any] = {}
    for field_number, excel_label, widget_key in METADATA_FIELD_SPECS:
        raw_value = st.session_state.get(widget_key)
        if field_number in NUMERIC_METADATA_FIELD_NUMBERS:
            parsed = _parse_required_number(
                _metadata_widget_state_to_string(field_number, raw_value)
            )
            if parsed is None:
                return (
                    {},
                    [],
                    [],
                    [],
                    [],
                    f"{field_number}. {excel_label} must be a number.",
                )
            metadata[excel_label] = _to_excel_number(parsed)
        elif field_number in DATE_METADATA_FIELD_NUMBERS:
            formatted = _format_metadata_date(raw_value)
            if not formatted:
                return (
                    {},
                    [],
                    [],
                    [],
                    [],
                    f"{field_number}. {excel_label} must be a valid date (YYYY-MM-DD).",
                )
            metadata[excel_label] = formatted
        elif field_number in TIME_METADATA_FIELD_NUMBERS:
            formatted = _format_metadata_time(raw_value)
            if not formatted:
                return (
                    {},
                    [],
                    [],
                    [],
                    [],
                    f"{field_number}. {excel_label} must be a valid time (e.g. 01:30 PM).",
                )
            metadata[excel_label] = formatted
        else:
            metadata[excel_label] = _metadata_widget_state_to_string(field_number, raw_value)

    target_values: list[int | float | str] = []
    for file, raw_target in zip(sorted_files, target_inputs, strict=True):
        stripped_target = raw_target.strip()
        if not stripped_target:
            target_values.append("")
            continue
        parsed_target = _parse_optional_number(raw_target)
        if parsed_target is None:
            return (
                {},
                [],
                [],
                [],
                [],
                (
                    f"{PREP_TARGET_CONCENTRATION_HEADER} for **{file.name}** must be a number "
                    "or left blank when there is no target CFU (e.g. heat kill / PAA panels)."
                ),
            )
        target_values.append(_to_excel_number(parsed_target))

    actual_values: list[int | float] = []
    for file, raw_actual in zip(sorted_files, actual_inputs, strict=True):
        parsed_actual = _parse_required_number(raw_actual)
        if parsed_actual is None:
            return (
                {},
                [],
                [],
                [],
                [],
                f"{PREP_ACTUAL_CONCENTRATION_HEADER} for **{file.name}** must be a number.",
            )
        actual_values.append(_to_excel_number(parsed_actual))

    source_txt_filenames = [file.name for file in sorted_files]
    treatment_inputs = special_treatment_inputs or [""] * len(sorted_files)
    special_treatments = [raw_treatment.strip() for raw_treatment in treatment_inputs]

    return metadata, target_values, actual_values, source_txt_filenames, special_treatments, None


def _build_embedded_workbook_rows(
    metadata: dict[str, Any],
    *,
    target_concentrations: list[int | float | str],
    actual_concentrations: list[int | float],
    source_txt_filenames: list[str],
    special_treatments: list[str],
    raman_shift: np.ndarray,
    intensity_columns: list[np.ndarray],
) -> list[list[Any]]:
    """
    Build row-major cell data for the embedded-metadata Excel layout.

    Parameters
    ----------
    metadata:
        Mapping of Excel column-A labels to column-B values (rows 1–16).
    target_concentrations:
        Target concentration labels per signal column.
    actual_concentrations:
        Measured concentration values per signal column.
    source_txt_filenames:
        Original instrument ``.txt`` filename per signal column (provenance).
    special_treatments:
        Optional treatment label per signal (e.g. heat kill, PAA); blank when unused.
    raman_shift:
        Raman shift grid (cm⁻¹).
    intensity_columns:
        One intensity vector per signal column.

    Returns
    -------
    list[list[Any]]
        Rectangular rows suitable for ``DataFrame.to_excel(header=False)``.
    """

    n_signals = len(intensity_columns)
    width = 1 + n_signals

    def _pad_row(row: list[Any]) -> list[Any]:
        return row + [""] * (width - len(row))

    rows: list[list[Any]] = [
        _pad_row([excel_label, metadata[excel_label]]) for _, excel_label, _ in METADATA_FIELD_SPECS
    ]
    rows.append(_pad_row([FILE_NAME_LABEL, *source_txt_filenames]))
    rows.append(_pad_row([SPECIAL_TREATMENT_LABEL, *special_treatments]))
    rows.append(_pad_row([TARGET_CONCENTRATION_LABEL, *target_concentrations]))
    rows.append(_pad_row([ACTUAL_CONCENTRATION_LABEL, *actual_concentrations]))
    rows.append([RAMAN_SHIFT_HEADER, *[INTENSITY_HEADER] * n_signals])

    for shift_idx, shift_value in enumerate(raman_shift):
        rows.append(
            [float(shift_value)]
            + [float(intensity_columns[signal_idx][shift_idx]) for signal_idx in range(n_signals)]
        )

    return rows


# Metadata field numbers (Excel row index) after which a full-width separator line is drawn.
_SEPARATOR_AFTER_METADATA_NUMBERS = frozenset(
    group_fields[-1] for group_fields in METADATA_LOGICAL_GROUPS
)


def _apply_full_width_row_separator(
    worksheet: Any,
    row_idx: int,
    max_col: int,
    bottom_side: Any,
) -> None:
    """
    Draw a horizontal separator beneath ``row_idx`` across columns 1..``max_col``.

    Only the bottom edge is styled; no left, right, or top borders are added.
    """
    from openpyxl.styles import Border

    for col_idx in range(1, max_col + 1):
        worksheet.cell(row=row_idx, column=col_idx).border = Border(bottom=bottom_side)


def _embedded_workbook_to_excel_bytes(rows: list[list[Any]]) -> bytes:
    """
    Serialize embedded workbook rows to styled ``.xlsx`` bytes.

    Applies bold labels and full-width horizontal separators between metadata
    sections, before the spectral table, and after the actual concentration row.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font, Side

    n_metadata = len(METADATA_FIELD_SPECS)
    n_signals = len(rows[n_metadata]) - 1
    data_width = 1 + n_signals
    file_name_row_idx = n_metadata + 1
    special_treatment_row_idx = n_metadata + 2
    target_row_idx = n_metadata + 3
    actual_row_idx = n_metadata + 4
    header_row_idx = n_metadata + 5

    separator_row_indices = [
        field_number
        for field_number, _, _ in METADATA_FIELD_SPECS
        if field_number in _SEPARATOR_AFTER_METADATA_NUMBERS
    ]
    separator_row_indices.append(actual_row_idx)

    thin_side = Side(style="thin", color="000000")
    bold_font = Font(bold=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            if value == "":
                continue
            worksheet.cell(row=row_idx, column=col_idx, value=value)

        if row_idx <= n_metadata:
            worksheet.cell(row=row_idx, column=1).font = bold_font
            continue

        if row_idx in {
            target_row_idx,
            actual_row_idx,
            file_name_row_idx,
            special_treatment_row_idx,
        }:
            worksheet.cell(row=row_idx, column=1).font = bold_font
            continue

        if row_idx == header_row_idx:
            for col_idx in range(1, data_width + 1):
                worksheet.cell(row=row_idx, column=col_idx).font = bold_font

    for separator_row_idx in separator_row_indices:
        _apply_full_width_row_separator(worksheet, separator_row_idx, data_width, thin_side)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _get_merged_preview() -> dict[str, Any] | None:
    """
    Return a valid merged-spectrum preview payload from session state.

    Clears stale values from older app versions (e.g. a legacy DataFrame).
    """

    payload = st.session_state.get(MERGED_PREVIEW_KEY)
    if payload is None:
        return None
    if isinstance(payload, pd.DataFrame) or not isinstance(payload, dict):
        st.session_state.pop(MERGED_PREVIEW_KEY, None)
        return None
    required_keys = {"raman_shift", "intensity_columns"}
    if not required_keys.issubset(payload.keys()):
        st.session_state.pop(MERGED_PREVIEW_KEY, None)
        return None
    n_signals = len(payload["intensity_columns"])
    has_concentrations = (
        "target_concentrations" in payload
        and "actual_concentrations" in payload
        and len(payload["target_concentrations"]) == n_signals
        and len(payload["actual_concentrations"]) == n_signals
    )
    has_legacy_labels = "labels" in payload and len(payload["labels"]) == n_signals
    if not has_concentrations and not has_legacy_labels:
        st.session_state.pop(MERGED_PREVIEW_KEY, None)
        return None
    return payload


def _format_merged_preview_legend_label(
    target: int | float | str,
    actual: int | float,
    *,
    source_txt_filename: str = "",
) -> str:
    """Format concentrations and source TXT name for the merged-signal preview legend."""

    target_display = target if str(target).strip() != "" else "—"
    base = f"Target: {target_display} / Actual: {actual}"
    if source_txt_filename:
        return f"{base} ({source_txt_filename})"
    return base


def _normalize_legacy_preview_label(label: str) -> str:
    """Upgrade older preview legend text to the colon-separated format."""

    match = re.match(
        r"Target[:\s]+([^/]+?)\s*/\s*Actual[:\s]+(.+)$",
        label.strip(),
        flags=re.IGNORECASE,
    )
    if match is None:
        return label
    return _format_merged_preview_legend_label(match.group(1).strip(), match.group(2).strip())


def _merged_preview_legend_labels(preview_payload: dict[str, Any]) -> list[str]:
    """Return legend labels for each merged signal column."""

    if "target_concentrations" in preview_payload and "actual_concentrations" in preview_payload:
        source_names = preview_payload.get("source_txt_filenames", [])
        return [
            _format_merged_preview_legend_label(
                target,
                actual,
                source_txt_filename=(source_names[idx] if idx < len(source_names) else ""),
            )
            for idx, (target, actual) in enumerate(
                zip(
                    preview_payload["target_concentrations"],
                    preview_payload["actual_concentrations"],
                    strict=True,
                )
            )
        ]
    return [_normalize_legacy_preview_label(label) for label in preview_payload["labels"]]


def _merged_preview_to_tidy_dataframe(preview_payload: dict[str, Any]) -> pd.DataFrame:
    """Convert merged preview arrays to the tidy schema required by ``plot_spectra``."""

    raman_shift = preview_payload["raman_shift"]
    legend_labels = _merged_preview_legend_labels(preview_payload)
    frames = [
        pd.DataFrame(
            {
                "raman_shift": raman_shift,
                "intensity": intensities,
                "filename": f"signal_{signal_index}",
                "signal_index": signal_index,
                MERGED_PREVIEW_HUE_COL: label,
            }
        )
        for signal_index, (intensities, label) in enumerate(
            zip(preview_payload["intensity_columns"], legend_labels, strict=True)
        )
    ]
    return pd.concat(frames, ignore_index=True)


def _render_merged_preview(preview_payload: dict[str, Any]) -> None:
    """Plot merged TXT spectra with the shared ``plot_spectra`` styling."""

    st.markdown("---")
    st.markdown("### Preview of Merged Signals")
    tidy_df = _merged_preview_to_tidy_dataframe(preview_payload)
    fig = plot_spectra(
        tidy_df,
        hue=MERGED_PREVIEW_HUE_COL,
        figsize=MERGED_PREVIEW_FIGSIZE,
        title="Merged Signal Preview",
    )
    legend = fig.axes[0].get_legend()
    if legend is not None:
        legend.set_title(MERGED_PREVIEW_LEGEND_TITLE)
        for text in legend.get_texts():
            text.set_fontsize(8)
    render_figure_stretch(fig)


@st.fragment
def _render_prep_export_and_preview() -> None:
    """
    Render post-convert download and spectrum preview in an isolated fragment.

    Keeps download clicks and preview redraws from rerunning metadata widgets,
    which reduces horizontal layout jitter on some browsers.
    """

    export_bytes = st.session_state.get(EXPORT_BYTES_KEY)
    export_filename = st.session_state.get(EXPORT_FILENAME_KEY)
    if export_bytes and export_filename:
        st.success("Merge complete. Download the embedded workbook below.")
        st.download_button(
            "Download Embedded Excel",
            data=export_bytes,
            file_name=str(export_filename),
            key="txt2excel_download_button",
            type="primary",
            on_click="ignore",
        )

    preview_payload = _get_merged_preview()
    if preview_payload is not None:
        _render_merged_preview(preview_payload)


def _process_uploaded_template(
    uploaded_template: Any,
) -> tuple[dict[str, Any], dict[str, str] | None]:
    """
    Parse a setup-template JSON upload and compute merged widget values.

    Returns
    -------
    tuple
        Feedback with ``level``, ``message``, and ``warnings``, plus merged widget
        values when parsing succeeds. Widget session state is not modified here;
        callers must defer application until before metadata widgets render.
    """

    try:
        payload = json.loads(uploaded_template.getvalue().decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        return {
            "level": "error",
            "message": f"Could not read template file: {exc}",
            "warnings": [],
        }, None

    if not isinstance(payload, dict):
        return {
            "level": "error",
            "message": "Template file must contain a JSON object.",
            "warnings": [],
        }, None

    before_values = _collect_metadata_values_by_widget_key()
    updated_values, warnings = _apply_template_import_to_values(payload, before_values)
    applied_count = sum(
        1
        for widget_key, value in updated_values.items()
        if value != before_values.get(widget_key, "")
    )
    if applied_count:
        message = f"Loaded template values for {applied_count} field(s)."
        level = "success"
    else:
        message = "Template applied; metadata already matched the file."
        level = "info"
    return {"level": level, "message": message, "warnings": warnings}, updated_values


def _render_template_import_feedback(container: DeltaGenerator) -> None:
    """Show import feedback stored from the prior apply-and-reset cycle."""

    feedback = st.session_state.pop(TEMPLATE_IMPORT_FEEDBACK_KEY, None)
    if not isinstance(feedback, dict):
        return
    level = feedback.get("level")
    message = feedback.get("message", "")
    warnings = feedback.get("warnings", [])
    if level == "success" and message:
        container.success(message)
    elif level == "info" and message:
        container.info(message)
    elif level == "error" and message:
        container.error(message)
    for warning in warnings:
        if warning:
            container.warning(warning)


def _render_export_template(container: DeltaGenerator) -> None:
    """Render selective export controls in a popover matching the import layout."""

    field_values = _collect_metadata_values_by_widget_key()
    with container.popover("Export metadata to template", use_container_width=True):
        st.caption(
            "Save metadata defaults to a JSON file. Only checked fields with valid "
            "values are included — you do not need to complete the form."
        )
        st.caption("Select fields to include in the download.")
        for field_number, _, widget_key in METADATA_FIELD_SPECS:
            selection_key = _template_selection_key(widget_key)
            if selection_key not in st.session_state:
                st.session_state[selection_key] = _is_template_field_exportable(
                    field_number,
                    field_values.get(widget_key, ""),
                )
        st.session_state[TEMPLATE_EXPORT_SELECT_ALL_KEY] = _all_template_export_fields_selected()
        st.checkbox(
            "Select all",
            key=TEMPLATE_EXPORT_SELECT_ALL_KEY,
            on_change=_apply_template_export_select_all,
        )
        n_columns = 2
        for row_start in range(0, len(METADATA_FIELD_SPECS), n_columns):
            row_specs = METADATA_FIELD_SPECS[row_start : row_start + n_columns]
            columns = st.columns(n_columns)
            for column, (field_number, excel_label, widget_key) in zip(
                columns, row_specs, strict=False
            ):
                selection_key = _template_selection_key(widget_key)
                column.checkbox(
                    f"{field_number}. {excel_label}",
                    key=selection_key,
                )

        selected_keys = frozenset(
            widget_key
            for _, _, widget_key in METADATA_FIELD_SPECS
            if st.session_state.get(_template_selection_key(widget_key), False)
        )
        export_payload = _build_template_export_payload(field_values, selected_keys)
        if export_payload is None:
            st.button(
                "Export metadata to template",
                disabled=True,
                help="Select at least one field with a valid value to export.",
                key="txt2excel_template_export_disabled",
                use_container_width=True,
            )
        else:
            st.download_button(
                "Export metadata to template",
                data=json.dumps(export_payload, indent=2),
                file_name=TEMPLATE_EXPORT_FILENAME,
                mime="application/json",
                key="txt2excel_template_export",
                use_container_width=True,
            )


def _render_import_template(container: DeltaGenerator) -> None:
    """
    Render import as a popover button matching the export template layout.

    ``st.file_uploader`` lives inside the popover because Streamlit does not
    expose a file-picker API on ``st.button``.
    """

    uploader_key = (
        f"txt2excel_template_import_"
        f"{st.session_state.get(TEMPLATE_IMPORT_UPLOADER_RESET_KEY, 'default')}"
    )
    with container.popover("Import metadata from template", use_container_width=True):
        st.caption(
            "Load a previously saved JSON file into the metadata form above. "
            "Every field stored in that file is applied — export field selection does not "
            "affect import."
        )
        st.caption("Select a `.json` setup file. Values apply as soon as the file is chosen.")
        uploaded_template = st.file_uploader(
            "JSON setup file",
            type=["json"],
            label_visibility="collapsed",
            key=uploader_key,
        )
        if uploaded_template is not None:
            feedback, updated_values = _process_uploaded_template(uploaded_template)
            st.session_state[TEMPLATE_IMPORT_FEEDBACK_KEY] = feedback
            if updated_values is not None:
                st.session_state[TEMPLATE_IMPORT_PENDING_VALUES_KEY] = updated_values
            st.session_state[TEMPLATE_IMPORT_UPLOADER_RESET_KEY] = str(uuid.uuid4())
            st.rerun()
    _render_template_import_feedback(container)


def _render_metadata_field_widget(
    column: DeltaGenerator,
    field_number: int,
    excel_label: str,
    widget_key: str,
    *,
    multiline: bool = False,
) -> None:
    """Render a single metadata widget using the appropriate Streamlit input type."""

    label = _metadata_field_label(field_number, excel_label)
    if field_number in DATE_METADATA_FIELD_NUMBERS:
        if widget_key not in st.session_state:
            st.session_state[widget_key] = None
        column.date_input(label, key=widget_key, format="YYYY-MM-DD")
        return
    if field_number in TIME_METADATA_FIELD_NUMBERS:
        if widget_key not in st.session_state:
            st.session_state[widget_key] = None
        column.time_input(label, key=widget_key)
        return
    if multiline:
        column.text_area(label, key=widget_key, height=100)
        return
    column.text_input(label, key=widget_key)


def _render_metadata_fields_in_columns(
    columns: list[DeltaGenerator],
    spec_by_number: dict[int, tuple[int, str, str]],
    field_numbers: tuple[int, ...],
    *,
    col_offset: int = 0,
) -> None:
    """
    Place metadata widgets into an existing column row starting at ``col_offset``.

    Parameters
    ----------
    columns:
        Column containers from ``st.columns(METADATA_UI_COLUMN_COUNT)``.
    spec_by_number:
        Mapping of field number to metadata spec triples.
    field_numbers:
        Field numbers to render in order.
    col_offset:
        Zero-based column index where the first field is placed.
    """

    for field_idx, field_number in enumerate(field_numbers):
        col_idx = col_offset + field_idx
        if col_idx >= len(columns):
            break
        number, excel_label, widget_key = spec_by_number[field_number]
        _render_metadata_field_widget(
            columns[col_idx],
            number,
            excel_label,
            widget_key,
        )


def _inject_prep_layout_stability_css() -> None:
    """
    Inject one-time CSS to reduce horizontal layout shift during Streamlit reruns.

    Reserving scrollbar gutter space helps prevent five-column metadata rows from
    pulsing wider/narrower when page height changes (common on Windows Chrome).
    """

    if st.session_state.get(_PREP_LAYOUT_STABILITY_CSS_KEY):
        return
    st.markdown(
        """
        <style>
        div[data-testid="stAppViewContainer"] {
            overflow-y: scroll;
        }
        div[data-testid="column"] {
            min-width: 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.session_state[_PREP_LAYOUT_STABILITY_CSS_KEY] = True


def _render_metadata_fields(container: DeltaGenerator) -> None:
    """
    Render numbered metadata in three five-column rows plus full-width notes.

    Row layout follows ``METADATA_UI_ROWS`` (5 + 5 + 5); Excel uses the same
    field order with dividers at each ``METADATA_LOGICAL_GROUPS`` boundary.
    """

    _inject_prep_layout_stability_css()
    container.markdown("### Experiment metadata")
    if st.session_state.pop(RESTORE_METADATA_AFTER_RELOAD_KEY, False):
        _restore_persistent_metadata_widgets()
    _apply_pending_template_import_values()
    spec_by_number = {
        field_number: (field_number, excel_label, widget_key)
        for field_number, excel_label, widget_key in METADATA_FIELD_SPECS
    }

    for row_fields in METADATA_UI_ROWS:
        row_columns = container.columns(METADATA_UI_COLUMN_COUNT)
        _render_metadata_fields_in_columns(row_columns, spec_by_number, row_fields)

    notes_field_number = METADATA_LOGICAL_GROUPS[-1][0]
    number, excel_label, widget_key = spec_by_number[notes_field_number]
    _render_metadata_field_widget(
        container,
        number,
        excel_label,
        widget_key,
        multiline=True,
    )

    _sync_persistent_metadata_snapshot()

    export_col, import_col = container.columns(2)
    _render_export_template(export_col)
    _render_import_template(import_col)


def _read_shift_bounds() -> tuple[str, str]:
    """Read min/max Raman shift widget values from session state."""

    min_shift = str(st.session_state.get("txt2excel_min_shift", DEFAULT_MIN_SHIFT)).strip()
    max_shift = str(st.session_state.get("txt2excel_max_shift", "")).strip()
    return min_shift, max_shift


def _sync_shift_defaults_for_upload(
    uploaded_files: list[Any],
    file_contents: dict[str, str],
) -> tuple[np.ndarray | None, str | None]:
    """
    Validate Raman shift grids across uploads and refresh default max-shift bounds.

    Returns
    -------
    tuple
        ``(common_shift, shift_error)``.
    """

    common_shift, shift_error = _validate_common_shift(file_contents)
    if shift_error or not uploaded_files:
        return common_shift, shift_error

    default_max_shift = str(common_shift.max()) if common_shift is not None else ""
    file_signature = tuple(sorted(file.name for file in uploaded_files))
    if st.session_state.get("txt2excel_file_signature") != file_signature:
        st.session_state["txt2excel_file_signature"] = file_signature
        if default_max_shift:
            st.session_state["txt2excel_max_shift"] = default_max_shift

    if "txt2excel_min_shift" not in st.session_state:
        st.session_state["txt2excel_min_shift"] = str(DEFAULT_MIN_SHIFT)

    return common_shift, shift_error


def _render_raman_shift_bounds(container: DeltaGenerator) -> None:
    """Render the Raman shift range subsection with min/max inputs."""

    container.markdown("### Raman shift range")
    col_min, col_max = container.columns(2)
    with col_min:
        col_min.text_input(
            _prep_field_label("Min Raman Shift"),
            key="txt2excel_min_shift",
        )
    with col_max:
        col_max.text_input(
            _prep_field_label("Max Raman Shift"),
            key="txt2excel_max_shift",
        )


def render_prep_mode() -> None:
    """Render the full TXT-to-Excel prep utility (separate app shell)."""

    with st.sidebar:
        st.button(
            "← Back to Analysis",
            key="exit_txt_to_excel_mode",
            on_click=enter_analysis_mode,
            use_container_width=True,
        )
        st.markdown("---")
        header_col, btn_col = st.columns([3, 1])
        with header_col:
            st.markdown("# 📁 Upload Raman TXT Files")
        with btn_col:
            st.button("Reload Data", type="primary", on_click=clear_prep_uploads)
        uploaded_files = st.file_uploader(
            "Upload Raman TXT Files",
            type=["txt"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key=f"txt_to_excel_uploader_{st.session_state.get(PREP_UPLOADER_RESET_KEY, 'default')}",
        )

    file_contents: dict[str, str] = {}
    if uploaded_files:
        for file in uploaded_files:
            file_contents[file.name] = file.read().decode("utf-8")

    shift_error: str | None = None
    if uploaded_files:
        _, shift_error = _sync_shift_defaults_for_upload(uploaded_files, file_contents)
        if shift_error:
            with st.sidebar:
                st.error(f"❌ {shift_error}")

    st.title("Raman TXT to Excel Merger")

    if not uploaded_files:
        st.info(
            "Upload one or more `.txt` files using the sidebar, then complete metadata "
            "required metadata (``*``), optional fields ``(optional)``, and per-signal rows below."
        )
        st.markdown("### Next steps")
        st.markdown(
            "1. Upload `.txt` files in the sidebar.\n"
            "2. Complete required metadata fields (marked with ``*``).\n"
            "3. Set the Raman shift range and per-signal concentrations (optional where noted).\n"
            "4. Convert and download the embedded `.xlsx`.\n"
            "5. Switch back to **Analysis** and upload the saved file."
        )
        return

    if shift_error:
        st.warning("Fix Raman shift mismatches before continuing.")
        return

    _render_metadata_fields(st)
    st.markdown("---")

    sorted_files = sorted(uploaded_files, key=lambda file: _extract_cfu_sort_key(file.name))

    _render_raman_shift_bounds(st)

    st.markdown("### Per-signal labels and concentrations")

    header_file, header_treatment, header_target, header_actual = st.columns([2, 1, 1, 1])
    with header_file:
        st.markdown("**File**")
    with header_treatment:
        st.markdown(_prep_column_header_markup(PREP_SPECIAL_TREATMENT_HEADER, optional=True))
    with header_target:
        st.markdown(_prep_column_header_markup(PREP_TARGET_CONCENTRATION_HEADER, optional=True))
    with header_actual:
        st.markdown(_prep_column_header_markup(PREP_ACTUAL_CONCENTRATION_HEADER))

    target_inputs: list[str] = []
    actual_inputs: list[str] = []
    special_treatment_inputs: list[str] = []
    for file in sorted_files:
        col_file, col_treatment, col_target, col_actual = st.columns([2, 1, 1, 1])
        with col_file:
            st.markdown(
                f"<div style='padding-top:6px'>{file.name}</div>",
                unsafe_allow_html=True,
            )
        with col_treatment:
            special_treatment_inputs.append(
                st.text_input(
                    label=f"Treatment for {file.name}",
                    key=f"txt2excel_treatment_{file.name}",
                    label_visibility="collapsed",
                    placeholder="e.g. heat kill",
                )
            )
        with col_target:
            target_inputs.append(
                st.text_input(
                    label=f"Target for {file.name}",
                    key=f"txt2excel_target_{file.name}",
                    label_visibility="collapsed",
                    placeholder="e.g. 1000",
                )
            )
        with col_actual:
            actual_inputs.append(
                st.text_input(
                    label=f"Actual for {file.name}",
                    key=f"txt2excel_actual_{file.name}",
                    label_visibility="collapsed",
                    placeholder="e.g. 995",
                )
            )

    st.markdown("---")
    st.markdown("### Output File Name")
    col_name_input, col_ext = st.columns([9, 1])
    with col_name_input:
        output_file_name = st.text_input(
            "Output File Name",
            placeholder="Enter output file name",
            key="txt2excel_output_name",
            label_visibility="collapsed",
        )
    with col_ext:
        st.markdown(
            "<div style='padding-top:6px;font-weight:600'>.xlsx</div>",
            unsafe_allow_html=True,
        )

    convert_clicked = st.button(
        "Convert and Export",
        key="txt2excel_convert_button",
        type="primary",
    )

    if convert_clicked:
        if not output_file_name.strip():
            st.warning("Please enter an output file name.")
            return

        (
            metadata,
            target_values,
            actual_values,
            source_txt_filenames,
            special_treatments,
            validation_error,
        ) = _validate_prep_inputs(
            sorted_files,
            target_inputs,
            actual_inputs,
            special_treatment_inputs=special_treatment_inputs,
        )
        if validation_error:
            st.error(validation_error)
            return

        min_shift_input, max_shift_input = _read_shift_bounds()
        try:
            min_shift = float(min_shift_input) if min_shift_input else None
            max_shift = float(max_shift_input) if max_shift_input else None
        except ValueError:
            st.error("Raman shift values must be numeric.")
            return

        ordered_contents = {file.name: file_contents[file.name] for file in sorted_files}
        raman_shift, intensity_columns, merge_error = _merge_txt_spectra(
            ordered_contents,
            min_shift=min_shift,
            max_shift=max_shift,
        )
        if merge_error or raman_shift is None or intensity_columns is None:
            st.error(f"❌ {merge_error or 'Merge failed.'}")
            st.session_state.pop(MERGED_PREVIEW_KEY, None)
            st.session_state.pop(EXPORT_BYTES_KEY, None)
            st.session_state.pop(EXPORT_FILENAME_KEY, None)
            return

        workbook_rows = _build_embedded_workbook_rows(
            metadata,
            target_concentrations=target_values,
            actual_concentrations=actual_values,
            source_txt_filenames=source_txt_filenames,
            special_treatments=special_treatments,
            raman_shift=raman_shift,
            intensity_columns=intensity_columns,
        )
        excel_bytes = _embedded_workbook_to_excel_bytes(workbook_rows)
        export_filename = f"{output_file_name.strip()}.xlsx"

        st.session_state[MERGED_PREVIEW_KEY] = {
            "raman_shift": raman_shift,
            "intensity_columns": intensity_columns,
            "target_concentrations": target_values,
            "actual_concentrations": actual_values,
            "source_txt_filenames": source_txt_filenames,
            "special_treatments": special_treatments,
        }
        st.session_state[EXPORT_BYTES_KEY] = excel_bytes
        st.session_state[EXPORT_FILENAME_KEY] = export_filename
        logger.info(
            "Embedded TXT merge complete: %d files → %d Raman rows, %d signal columns",
            len(sorted_files),
            len(raman_shift),
            len(intensity_columns),
        )

    _render_prep_export_and_preview()
